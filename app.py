"""
GitHub Issue Cost Estimator - Flask Backend
Main application file handling API endpoints and business logic
"""

import os
import re
import json
import uuid
import threading
from io import StringIO, BytesIO
from typing import List, Dict, Optional
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import html

from llm_analyzer import LLMAnalyzer

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# Initialize LLM analyzer
llm_analyzer = LLMAnalyzer()

# Store results in memory (in production, use a database)
analysis_cache = {}

# Store progress tracking for each analysis session
progress_tracking = {}


class GitHubAPIClient:
    """Handles all GitHub API interactions"""

    BASE_URL = "https://api.github.com"

    def __init__(self):
        self.token = os.getenv('GITHUB_TOKEN')
        self.headers = {
            'Accept': 'application/vnd.github.v3+json'
        }
        if self.token:
            self.headers['Authorization'] = f'token {self.token}'
            print(f"GitHub token loaded: {self.token[:8]}...{self.token[-4:] if len(self.token) > 12 else 'short'}")
        else:
            print("WARNING: No GitHub token found - using unauthenticated requests (60/hour limit)")

    def parse_repo_url(self, url: str) -> tuple:
        """
        Parse GitHub repository URL to extract owner and repo name

        Args:
            url: GitHub repository URL

        Returns:
            tuple: (owner, repo) or raises ValueError
        """
        patterns = [
            r'github\.com/([^/]+)/([^/]+)',
            r'^([^/]+)/([^/]+)$'
        ]

        for pattern in patterns:
            match = re.search(pattern, url.strip().rstrip('/'))
            if match:
                owner, repo = match.groups()
                # Remove .git suffix if present
                repo = repo.replace('.git', '')
                return owner, repo

        raise ValueError("Invalid GitHub repository URL format")

    def fetch_issues(self, owner: str, repo: str) -> List[Dict]:
        """
        Fetch all open issues from a GitHub repository with pagination

        Args:
            owner: Repository owner
            repo: Repository name

        Returns:
            List of issue dictionaries
        """
        issues = []
        page = 1
        per_page = 100

        print(f"Starting to fetch issues from {owner}/{repo}")

        while True:
            url = f"{self.BASE_URL}/repos/{owner}/{repo}/issues"
            params = {
                'state': 'open',
                'per_page': per_page,
                'page': page
            }

            print(f"Fetching page {page} (up to {per_page} items per page)")
            response = requests.get(url, headers=self.headers, params=params)
            
            # Debug rate limit info
            rate_limit_remaining = response.headers.get('X-RateLimit-Remaining', 'unknown')
            rate_limit_limit = response.headers.get('X-RateLimit-Limit', 'unknown')
            print(f"Rate limit: {rate_limit_remaining}/{rate_limit_limit} remaining")

            if response.status_code == 404:
                raise ValueError("Repository not found or is private")
            elif response.status_code == 403:
                # Check rate limit headers
                rate_limit_remaining = response.headers.get('X-RateLimit-Remaining', 'unknown')
                rate_limit_reset = response.headers.get('X-RateLimit-Reset', 'unknown')
                print(f"Rate limit hit. Remaining: {rate_limit_remaining}, Reset: {rate_limit_reset}")
                print(f"Headers: {dict(response.headers)}")
                
                if rate_limit_remaining == '0':
                    raise ValueError(f"GitHub API rate limit exceeded (0 requests remaining). Please add a GITHUB_TOKEN environment variable for higher limits (5000/hour vs 60/hour). Current issues fetched: {len(issues)}")
                else:
                    raise ValueError("API rate limit exceeded. Please provide a GitHub token.")
            elif response.status_code != 200:
                print(f"API Error {response.status_code}: {response.text}")
                raise ValueError(f"GitHub API error: {response.status_code}")

            batch = response.json()
            print(f"Page {page}: Received {len(batch)} total items from API")

            # Filter out pull requests (GitHub API returns PRs as issues)
            issues_only = [issue for issue in batch if 'pull_request' not in issue]
            prs_filtered = len(batch) - len(issues_only)
            
            print(f"Page {page}: {len(issues_only)} issues, {prs_filtered} PRs filtered out")

            if not issues_only:
                print(f"No more issues found on page {page}, stopping pagination")
                break

            issues.extend(issues_only)
            print(f"Total issues collected so far: {len(issues)}")

            # Check if there are more pages - use original batch size, not filtered
            if len(batch) < per_page:
                print(f"Last page reached (got {len(batch)} < {per_page})")
                break

            page += 1

        print(f"Finished fetching. Total issues: {len(issues)}")
        return issues

    def extract_issue_data(self, issue: Dict) -> Dict:
        """
        Extract relevant data from GitHub issue object

        Args:
            issue: Raw GitHub issue object

        Returns:
            Cleaned issue dictionary
        """
        return {
            'issue_number': issue['number'],
            'title': issue['title'],
            'body': issue['body'] or '',
            'labels': [label['name'] for label in issue.get('labels', [])],
            'html_url': issue['html_url'],
            'created_at': issue['created_at'],
            'updated_at': issue['updated_at']
        }


@app.route('/')
def index():
    """Serve the main HTML page"""
    return app.send_static_file('index.html')


def process_single_repo(repo_url, hourly_rate, github_client, repo_index, total_repos, session_id):
    """
    Process a single repository and return results

    Args:
        repo_url: GitHub repository URL
        hourly_rate: Engineer cost per hour
        github_client: GitHubAPIClient instance
        repo_index: Index of this repo (1-based)
        total_repos: Total number of repos being analyzed
        session_id: Session ID for progress tracking

    Returns:
        Dictionary with analysis results or error
    """
    try:
        # Parse repository URL
        owner, repo = github_client.parse_repo_url(repo_url)

        # Fetch issues from GitHub
        print(f"[Repo {repo_index}/{total_repos}] Fetching issues from {owner}/{repo}...")
        raw_issues = github_client.fetch_issues(owner, repo)

        if not raw_issues:
            return {
                'repo_url': repo_url,
                'owner': owner,
                'repo': repo,
                'issues': [],
                'total_cost': 0,
                'total_hours': 0,
                'issue_count': 0,
                'status': 'success',
                'message': 'No open issues found'
            }

        print(f"[Repo {repo_index}/{total_repos}] Found {len(raw_issues)} open issues")

        # Extract and clean issue data
        issues = [github_client.extract_issue_data(issue) for issue in raw_issues]

        # Analyze each issue with LLM
        analyzed_issues = []
        total_cost = 0

        for idx, issue in enumerate(issues, 1):
            print(f"[Repo {repo_index}/{total_repos}] Analyzing issue {idx}/{len(issues)}: #{issue['issue_number']}")

            # Update progress for this specific issue
            base_progress = int((repo_index - 1) / total_repos * 100)
            repo_progress = int((idx / len(issues)) * (100 / total_repos))
            total_progress = min(base_progress + repo_progress, 99)

            progress_tracking[session_id].update({
                'progress': total_progress,
                'message': f'Repo {repo_index}/{total_repos}: Analyzing issue {idx}/{len(issues)}: {issue["title"][:50]}...',
                'current_repo': repo_index,
                'current_issue': idx,
                'total_issues_in_repo': len(issues)
            })

            try:
                analysis = llm_analyzer.analyze_issue(
                    title=issue['title'],
                    body=issue['body'],
                    labels=issue['labels']
                )

                # Calculate cost based on hours and hourly rate
                estimated_hours = analysis['estimated_hours']
                estimated_cost = estimated_hours * hourly_rate

                analyzed_issue = {
                    'issue_number': issue['issue_number'],
                    'title': issue['title'],
                    'complexity': analysis['complexity'],
                    'estimated_hours': estimated_hours,
                    'estimated_cost': round(estimated_cost, 2),
                    'labels': ', '.join(issue['labels']),
                    'url': issue['html_url'],
                    'reasoning': analysis['reasoning']
                }

                analyzed_issues.append(analyzed_issue)
                total_cost += estimated_cost

            except Exception as e:
                print(f"Error analyzing issue #{issue['issue_number']}: {str(e)}")
                analyzed_issues.append({
                    'issue_number': issue['issue_number'],
                    'title': issue['title'],
                    'complexity': 'Unknown',
                    'estimated_hours': 0,
                    'estimated_cost': 0,
                    'labels': ', '.join(issue['labels']),
                    'url': issue['html_url'],
                    'reasoning': 'Analysis failed. Manual review required.'
                })

        # Cache results for CSV download
        cache_key = f"{owner}_{repo}"
        analysis_cache[cache_key] = analyzed_issues

        # Calculate total hours
        total_hours = sum(issue.get('estimated_hours', 0) for issue in analyzed_issues)

        return {
            'repo_url': repo_url,
            'owner': owner,
            'repo': repo,
            'issues': analyzed_issues,
            'total_cost': round(total_cost, 2),
            'total_hours': round(total_hours, 1),
            'issue_count': len(analyzed_issues),
            'cache_key': cache_key,
            'status': 'success'
        }

    except Exception as e:
        print(f"[Repo {repo_index}/{total_repos}] Error: {str(e)}")
        return {
            'repo_url': repo_url,
            'owner': '',
            'repo': '',
            'issues': [],
            'total_cost': 0,
            'total_hours': 0,
            'issue_count': 0,
            'status': 'error',
            'error': str(e)
        }


def process_multiple_repos(session_id, repo_urls, hourly_rate):
    """
    Background function to process multiple repositories
    """
    try:
        total_repos = len(repo_urls)
        github_client = GitHubAPIClient()
        repo_results = []

        for repo_index, repo_url in enumerate(repo_urls, 1):
            # Update progress for current repo
            progress_percent = int((repo_index - 1) / total_repos * 100)
            progress_tracking[session_id].update({
                'progress': progress_percent,
                'message': f'Analyzing repository {repo_index}/{total_repos}...',
                'current_repo': repo_index,
                'total_repos': total_repos
            })

            # Process this repository
            result = process_single_repo(repo_url, hourly_rate, github_client, repo_index, total_repos, session_id)
            repo_results.append(result)

            # Update progress tracking with intermediate results
            progress_tracking[session_id]['repo_results'] = repo_results

        # Calculate overall totals
        total_cost = sum(r.get('total_cost', 0) for r in repo_results)
        total_hours = sum(r.get('total_hours', 0) for r in repo_results)
        total_issues = sum(r.get('issue_count', 0) for r in repo_results)

        # Mark as complete with all results
        progress_tracking[session_id].update({
            'status': 'complete',
            'progress': 100,
            'message': 'Analysis complete!',
            'result': {
                'repo_results': repo_results,
                'total_cost': round(total_cost, 2),
                'total_hours': round(total_hours, 1),
                'total_issues': total_issues,
                'hourly_rate': hourly_rate,
                'session_id': session_id
            }
        })

    except Exception as e:
        print(f"Error in background analysis: {str(e)}")
        progress_tracking[session_id].update({
            'status': 'error',
            'progress': 0,
            'message': f'Error: {str(e)}'
        })


def process_analysis(session_id, repo_url, hourly_rate):
    """
    Background function to process analysis (legacy single-repo support)
    """
    try:
        progress_tracking[session_id].update({
            'progress': 5,
            'message': f'Parsing repository URL...'
        })

        # Initialize GitHub client
        github_client = GitHubAPIClient()

        # Parse repository URL
        try:
            owner, repo = github_client.parse_repo_url(repo_url)
            progress_tracking[session_id].update({
                'progress': 10,
                'message': f'Fetching issues from {owner}/{repo}...'
            })
        except ValueError as e:
            progress_tracking[session_id].update({
                'status': 'error',
                'progress': 0,
                'message': str(e)
            })
            return

        # Fetch issues from GitHub
        print(f"Fetching issues from {owner}/{repo}...")
        raw_issues = github_client.fetch_issues(owner, repo)

        if not raw_issues:
            progress_tracking[session_id].update({
                'status': 'complete',
                'progress': 100,
                'message': 'No open issues found',
                'result': {
                    'message': 'No open issues found in this repository',
                    'issues': [],
                    'total_cost': 0,
                    'repo_info': {'owner': owner, 'repo': repo}
                }
            })
            return

        print(f"Found {len(raw_issues)} open issues")

        # Update progress tracking
        progress_tracking[session_id].update({
            'status': 'fetching',
            'progress': 15,
            'message': f'Found {len(raw_issues)} open issues',
            'total': len(raw_issues),
            'current': 0
        })

        # Extract and clean issue data
        issues = [github_client.extract_issue_data(issue) for issue in raw_issues]

        progress_tracking[session_id].update({
            'status': 'analyzing',
            'progress': 20,
            'message': 'Starting AI analysis...'
        })

        # Analyze each issue with LLM
        analyzed_issues = []
        total_cost = 0

        for idx, issue in enumerate(issues, 1):
            print(f"Analyzing issue {idx}/{len(issues)}: #{issue['issue_number']}")

            # Update progress
            progress_percent = 20 + int((idx / len(issues)) * 75)  # 20% to 95%
            progress_tracking[session_id].update({
                'progress': progress_percent,
                'current': idx,
                'message': f'Analyzing issue {idx}/{len(issues)}: {issue["title"][:50]}...'
            })

            try:
                analysis = llm_analyzer.analyze_issue(
                    title=issue['title'],
                    body=issue['body'],
                    labels=issue['labels']
                )

                # Calculate cost based on hours and hourly rate
                estimated_hours = analysis['estimated_hours']
                estimated_cost = estimated_hours * hourly_rate

                analyzed_issue = {
                    'issue_number': issue['issue_number'],
                    'title': issue['title'],
                    'complexity': analysis['complexity'],
                    'estimated_hours': estimated_hours,
                    'estimated_cost': round(estimated_cost, 2),
                    'labels': ', '.join(issue['labels']),
                    'url': issue['html_url'],
                    'reasoning': analysis['reasoning']
                }

                analyzed_issues.append(analyzed_issue)
                total_cost += estimated_cost

            except Exception as e:
                print(f"Error analyzing issue #{issue['issue_number']}: {str(e)}")
                # Add with default values if analysis fails
                analyzed_issues.append({
                    'issue_number': issue['issue_number'],
                    'title': issue['title'],
                    'complexity': 'Unknown',
                    'estimated_hours': 0,
                    'estimated_cost': 0,
                    'labels': ', '.join(issue['labels']),
                    'url': issue['html_url'],
                    'reasoning': 'Analysis failed. Manual review required.'
                })

        # Cache results for CSV download
        cache_key = f"{owner}_{repo}"
        analysis_cache[cache_key] = analyzed_issues

        # Calculate total hours
        total_hours = sum(issue.get('estimated_hours', 0) for issue in analyzed_issues)

        # Mark as complete with results
        progress_tracking[session_id].update({
            'status': 'complete',
            'progress': 100,
            'message': 'Analysis complete!',
            'result': {
                'issues': analyzed_issues,
                'total_cost': round(total_cost, 2),
                'total_hours': round(total_hours, 1),
                'hourly_rate': hourly_rate,
                'repo_info': {
                    'owner': owner,
                    'repo': repo,
                    'issue_count': len(analyzed_issues)
                },
                'cache_key': cache_key,
                'session_id': session_id
            }
        })

    except Exception as e:
        print(f"Error in background analysis: {str(e)}")
        progress_tracking[session_id].update({
            'status': 'error',
            'progress': 0,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/analyze', methods=['POST'])
def analyze_issues():
    """
    Main endpoint to analyze GitHub repository issues - returns immediately with session_id

    Expected JSON body:
        {
            "repo_urls": ["https://github.com/owner/repo1", "https://github.com/owner/repo2"],
            "hourly_rate": 80
        }

    Returns:
        JSON with session_id for progress tracking
    """
    try:
        data = request.get_json()
        repo_urls = data.get('repo_urls', [])
        hourly_rate = float(data.get('hourly_rate', 80))  # Default $80/hour

        # Support both single and multiple repos for backward compatibility
        if not repo_urls and data.get('repo_url'):
            repo_urls = [data.get('repo_url')]

        # Filter out empty URLs
        repo_urls = [url.strip() for url in repo_urls if url and url.strip()]

        if not repo_urls:
            return jsonify({'error': 'At least one repository URL is required'}), 400

        if len(repo_urls) > 5:
            return jsonify({'error': 'Maximum 5 repositories allowed'}), 400

        if hourly_rate <= 0:
            return jsonify({'error': 'Hourly rate must be positive'}), 400

        # Generate session ID early for progress tracking
        session_id = str(uuid.uuid4())

        # Initialize progress tracking - fetching stage
        progress_tracking[session_id] = {
            'status': 'connecting',
            'progress': 0,
            'message': 'Connecting to GitHub...',
            'total': 0,
            'current': 0,
            'result': None,
            'repo_count': len(repo_urls),
            'repo_results': []
        }

        # Start background processing for multiple repos
        thread = threading.Thread(
            target=process_multiple_repos,
            args=(session_id, repo_urls, hourly_rate)
        )
        thread.daemon = True
        thread.start()

        # Return immediately with session_id
        return jsonify({
            'session_id': session_id,
            'message': 'Analysis started',
            'repo_count': len(repo_urls)
        }), 202

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/progress/<session_id>', methods=['GET'])
def get_progress(session_id):
    """
    Get progress for an analysis session

    Args:
        session_id: Session ID from analyze endpoint

    Returns:
        JSON with progress information
    """
    if session_id in progress_tracking:
        return jsonify(progress_tracking[session_id])
    else:
        return jsonify({
            'status': 'not_found',
            'progress': 0,
            'message': 'Session not found'
        }), 404


@app.route('/api/download-csv', methods=['POST'])
def download_csv():
    """
    Generate and download Excel file of analyzed issues with proper formatting

    Expected JSON body:
        {
            "cache_key": "owner_repo",
            "issues": [...] (optional, if not using cache)
        }
    """
    try:
        data = request.get_json()
        cache_key = data.get('cache_key')
        issues = data.get('issues')

        # Get issues from cache or request body
        if cache_key and cache_key in analysis_cache:
            issues = analysis_cache[cache_key]

        if not issues:
            return jsonify({'error': 'No data available for Excel generation'}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Issue Analysis"

        # Define headers with proper naming
        headers = ['Issue #', 'Title', 'Complexity', 'Hours', 'Cost', 'Labels', 'URL', 'Reasoning']

        # Add headers to first row
        ws.append(headers)

        # Style headers: Bold, Black background, White text
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for issue in issues:
            # Strip HTML tags from reasoning
            reasoning_text = issue.get('reasoning', '')
            if reasoning_text:
                # Remove HTML tags and convert to plain text
                reasoning_text = html.unescape(re.sub('<[^<]+?>', '', reasoning_text))
                # Replace multiple spaces/newlines with single space
                reasoning_text = ' '.join(reasoning_text.split())

            row_data = [
                issue.get('issue_number', ''),
                issue.get('title', ''),
                issue.get('complexity', ''),
                issue.get('estimated_hours', 0),
                issue.get('estimated_cost', 0),
                issue.get('labels', ''),
                issue.get('url', ''),
                reasoning_text
            ]
            ws.append(row_data)

        # Adjust column widths
        column_widths = {
            'A': 10,   # Issue #
            'B': 50,   # Title
            'C': 12,   # Complexity
            'D': 10,   # Hours
            'E': 12,   # Cost
            'F': 30,   # Labels
            'G': 50,   # URL
            'H': 80    # Reasoning
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Set row height for header
        ws.row_dimensions[1].height = 25

        # Align data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        # Send file
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='issue_analysis.xlsx'
        )

    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'llm_provider': os.getenv('LLM_PROVIDER', 'not configured')
    })


if __name__ == '__main__':
    # Check for API keys
    llm_provider = os.getenv('LLM_PROVIDER', 'anthropic')
    if llm_provider == 'anthropic' and not os.getenv('ANTHROPIC_API_KEY'):
        print("WARNING: ANTHROPIC_API_KEY not set in .env file")
    elif llm_provider == 'openai' and not os.getenv('OPENAI_API_KEY'):
        print("WARNING: OPENAI_API_KEY not set in .env file")

    print(f"Using LLM provider: {llm_provider}")
    
    # Production configuration
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    
    print(f"Starting Flask server on port {port}")
    app.run(host='0.0.0.0', port=port, debug=debug)
